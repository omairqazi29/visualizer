"""Parser for DHS Yearbook Tables 8-11 "New Arrivals / Adjustments of Status".

These workbooks break each Yearbook table into the two ways a person can obtain
LPR status:

- **New arrivals** = admitted on an immigrant visa issued abroad (consular).
- **Adjustments of status** = adjusted from inside the US (I-485 / AOS).

Table 10 ("Persons Obtaining Lawful Permanent Resident Status by Region and
Country of Birth") carries an "Employment-based preferences" column, which makes
these files the only source for the EB consular-vs-AOS split by fiscal year and
by country of birth.

Why this matters: DOS issuance data is consular-only, and USCIS I-485 data is
AOS-only. Neither alone gives total EB usage. `src/engine/supply.py` needs the
total (see AGENTS.md, Data-Driven Supply Model rule 2), and these files supply
the aos/consular columns of `dhs_eb_category_usage.csv`.

Two layouts exist and both are handled:

1. **Combined** (FY2018-FY2022): a single Table 10 sheet with two side-by-side
   blocks. Row 3 labels the blocks ("Adjustments of status", "New arrivals");
   row 4 names the columns within each block.
2. **Split** (FY2023+): separate "Table 10 New Arrivals" and "Table 10 Adjust"
   sheets, each with its own header row.

Note on precision: FY2023+ releases are rounded to the nearest 10 by DHS.
Earlier releases carry exact counts. Values are reported as published.

Granularity limit: Table 10 reports employment-based as a single total, not
split by preference category. Per-category (EB1-EB5) figures require Table 7
of the main Yearbook workbook, handled by `dhs_yearbook_parser.py`.
"""

from pathlib import Path
from typing import Optional
import re

import openpyxl


__all__ = ["DHSNewAdjParser", "EBSplit"]


# Marker text used to locate structure within the sheets (all matched casefolded).
_ROW_LABEL_HEADER = "region and country of birth"
_EB_COLUMN = "employment-based"
_TOTAL_COLUMN = "total"
# "adjust" (not "adjustment") so this matches both the combined-layout block
# label "Adjustments of status" and the FY2023+ sheet named "Table 10 Adjust".
_ADJUST_BLOCK = "adjust"
_ARRIVALS_BLOCK = "new arrival"

# Sheet-name matching
_TABLE10 = "table 10"

# DHS disclosure placeholders: D = withheld for confidentiality, - = zero/none.
_SUPPRESSED = {"D", "-", "", "X", "NA", "N/A"}

# Yearbook country labels differ from the short names used elsewhere in this
# project (constants.py, the DOS parser). Each key maps to the labels that may
# appear in the "Region and country of birth" column across releases.
_COUNTRY_ALIASES: dict[str, tuple[str, ...]] = {
    "china": (
        "china, people's republic",
        "china, people's republic of",
        "china, mainland",
        "china",
    ),
    "south korea": ("korea, south", "korea, republic of", "south korea"),
    "vietnam": ("vietnam", "viet nam"),
    "russia": ("russia", "russian federation"),
    "burma": ("burma", "myanmar"),
}


def _country_candidates(country: str) -> tuple[str, ...]:
    """Return the cleaned labels to look for when matching a country row."""
    key = _clean(country)
    aliases = _COUNTRY_ALIASES.get(key)
    if aliases:
        # Keep the caller's own spelling in play alongside the known aliases
        return tuple(dict.fromkeys((*aliases, key)))
    return (key,)


class EBSplit:
    """Employment-based LPR counts for one fiscal year, split by immigration path."""

    def __init__(self, fiscal_year: int, aos: Optional[int], consular: Optional[int], source_file: str):
        self.fiscal_year = fiscal_year
        self.aos = aos
        self.consular = consular
        self.source_file = source_file

    @property
    def total(self) -> Optional[int]:
        """Combined EB usage, or None when either path is unavailable."""
        if self.aos is None or self.consular is None:
            return None
        return self.aos + self.consular

    def as_dict(self) -> dict:
        return {
            "fiscal_year": self.fiscal_year,
            "total": self.total,
            "aos": self.aos,
            "consular": self.consular,
            "source_file": self.source_file,
        }

    def __repr__(self) -> str:
        return (
            f"EBSplit(fy={self.fiscal_year}, total={self.total}, "
            f"aos={self.aos}, consular={self.consular})"
        )


def _clean(value) -> str:
    """Normalize a cell to a comparable lowercase string."""
    if value is None:
        return ""
    return " ".join(str(value).split()).casefold()


def _to_int(value) -> Optional[int]:
    """Convert a Table 10 count cell to int, or None when suppressed/unparseable.

    Handles the 'D' (withheld) and '-' (none) disclosure markers that appear
    throughout government workbooks, plus comma-formatted strings.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        # Excel stores these as numbers in some releases, strings in others
        return int(round(float(value)))
    text = str(value).strip().replace(",", "")
    if text.upper() in _SUPPRESSED:
        return None
    try:
        return int(round(float(text)))
    except ValueError:
        return None


def _parse_fy(name: str) -> Optional[int]:
    """Extract the fiscal year a tables8-11 workbook covers from its filename.

    The publication year in the filename prefix (e.g. '2026_0604_ohss_...') is
    deliberately ignored: only the explicit fy#### token identifies the data year.
    """
    m = re.search(r"fy[_-]?(\d{4})", name, re.IGNORECASE)
    if m:
        return int(m.group(1))
    return None


class DHSNewAdjParser:
    """Reads EB new-arrival vs adjustment splits from DHS Tables 8-11 workbooks."""

    def __init__(self, data_dir: str = "data/DHS_Yearbook"):
        self.data_dir = Path(data_dir)
        self._cache: Optional[dict[int, EBSplit]] = None

    # ──────────────────────────────────────────────
    # File discovery
    # ──────────────────────────────────────────────

    def _discover_files(self) -> dict[int, Path]:
        """Map fiscal year to its tables8-11 workbook.

        Government filenames vary widely across releases (fy2018_tables8-11newadj_d,
        plcy_tables8-11newadj_fy2022_d, 2026_0604_ohss_tables8-11newadj_fy2024).
        When several files claim the same fiscal year, the most recently modified
        wins, since DHS reissues corrected versions under new names.
        """
        found: dict[int, Path] = {}
        if not self.data_dir.is_dir():
            return found
        for path in sorted(self.data_dir.glob("*tables8-11*.xlsx")):
            fy = _parse_fy(path.name)
            if fy is None:
                continue
            existing = found.get(fy)
            if existing is None or path.stat().st_mtime > existing.stat().st_mtime:
                found[fy] = path
        return found

    # ──────────────────────────────────────────────
    # Sheet-level extraction
    # ──────────────────────────────────────────────

    @staticmethod
    def _table10_sheets(wb) -> list[str]:
        return [s for s in wb.sheetnames if _TABLE10 in _clean(s)]

    @staticmethod
    def _read_grid(ws, max_row: int = 250) -> list[list]:
        """Materialize a sheet into memory.

        Read-only worksheets are forward-only streams in openpyxl, so a sheet
        can only be iterated once. Everything downstream (header detection,
        column location, row lookup) reads this single materialized copy.
        """
        return [list(r) for r in ws.iter_rows(max_row=max_row, values_only=True)]

    @classmethod
    def _find_header(cls, grid: list[list]) -> Optional[int]:
        """Return the 0-based index of the row labelling the row-header column.

        Matched exactly rather than by substring: the sheet's title row reads
        "PERSONS OBTAINING LAWFUL PERMANENT RESIDENT STATUS BY REGION AND
        COUNTRY OF BIRTH...", which contains the header text and would
        otherwise win by appearing first.
        """
        for i, row in enumerate(grid):
            if not row:
                continue
            # Trailing footnote markers appear on some releases, e.g. "... birth 1"
            label = _clean(row[0]).rstrip(" 0123456789").strip()
            if label == _ROW_LABEL_HEADER:
                return i
        return None

    @classmethod
    def _eb_column_split_layout(cls, grid: list[list], header_idx: int) -> Optional[int]:
        """Column index of 'Employment-based preferences' in a split-sheet layout."""
        header = grid[header_idx]
        for col, cell in enumerate(header):
            if _EB_COLUMN in _clean(cell):
                return col
        return None

    @classmethod
    def _eb_columns_combined_layout(cls, grid: list[list], header_idx: int) -> tuple[Optional[int], Optional[int]]:
        """Return (aos_eb_col, consular_eb_col) for the side-by-side layout.

        Row `header_idx` labels the two blocks and leaves the rest of each block
        blank (merged cells read as None), so block extent is derived by scanning
        forward from each label. Row `header_idx + 1` names the columns.
        """
        block_row = grid[header_idx]
        name_row = grid[header_idx + 1] if header_idx + 1 < len(grid) else []
        if not name_row:
            return (None, None)

        # Locate where each block starts
        starts: list[tuple[int, str]] = []
        for col, cell in enumerate(block_row):
            text = _clean(cell)
            if _ADJUST_BLOCK in text:
                starts.append((col, "aos"))
            elif _ARRIVALS_BLOCK in text:
                starts.append((col, "consular"))
        if not starts:
            return (None, None)
        starts.sort()

        result = {"aos": None, "consular": None}
        for i, (start, kind) in enumerate(starts):
            end = starts[i + 1][0] if i + 1 < len(starts) else len(name_row)
            for col in range(start, min(end, len(name_row))):
                if _EB_COLUMN in _clean(name_row[col]):
                    result[kind] = col
                    break
        return (result["aos"], result["consular"])

    @staticmethod
    def _total_row_value(grid: list[list], label_col: int, value_col: int) -> Optional[int]:
        """Read the national 'Total' row for one column.

        Table 10 repeats a 'Total' row under both the REGION and COUNTRY blocks
        with identical values; the first parseable one is taken.
        """
        for row in grid:
            if not row or label_col >= len(row):
                continue
            if _clean(row[label_col]) == _TOTAL_COLUMN:
                if value_col < len(row):
                    value = _to_int(row[value_col])
                    if value is not None:
                        return value
        return None

    @staticmethod
    def _country_row_value(grid: list[list], label_col: int, value_col: int, country: str) -> Optional[int]:
        """Read a specific country row for one column.

        Table 10 lists each country twice (once under REGION's subtotals is not
        applicable, but the COUNTRY block repeats for both halves of some
        releases); the first match is authoritative.
        """
        targets = _country_candidates(country)
        for row in grid:
            if not row or label_col >= len(row):
                continue
            label = _clean(row[label_col])
            # Yearbook country labels carry footnote markers, e.g. "China 1"
            stripped = label.rstrip(" 0123456789").strip()
            if label in targets or stripped in targets:
                if value_col < len(row):
                    return _to_int(row[value_col])
        return None

    def _extract(self, path: Path, country: Optional[str] = None) -> tuple[Optional[int], Optional[int]]:
        """Return (aos, consular) EB counts from one workbook.

        When `country` is given the figures are for that country of birth,
        otherwise they are the national totals.

        An unreadable or truncated workbook yields (None, None) rather than
        raising, so one bad file in the directory cannot break the whole series.
        """
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            return (None, None)
        try:
            sheets = self._table10_sheets(wb)
            if not sheets:
                return (None, None)

            arrivals = [s for s in sheets if _ARRIVALS_BLOCK in _clean(s) and _ADJUST_BLOCK not in _clean(s)]
            adjust = [s for s in sheets if _ADJUST_BLOCK in _clean(s) and _ARRIVALS_BLOCK not in _clean(s)]

            def read(grid: list[list], col: Optional[int]) -> Optional[int]:
                if col is None:
                    return None
                if country is None:
                    return self._total_row_value(grid, 0, col)
                return self._country_row_value(grid, 0, col, country)

            # Split layout: one sheet per path
            if arrivals and adjust:
                values = {}
                for kind, name in (("consular", arrivals[0]), ("aos", adjust[0])):
                    grid = self._read_grid(wb[name])
                    header_idx = self._find_header(grid)
                    if header_idx is None:
                        values[kind] = None
                        continue
                    col = self._eb_column_split_layout(grid, header_idx)
                    values[kind] = read(grid, col)
                return (values.get("aos"), values.get("consular"))

            # Combined layout: both blocks on one sheet
            grid = self._read_grid(wb[sheets[0]])
            header_idx = self._find_header(grid)
            if header_idx is None:
                return (None, None)
            aos_col, consular_col = self._eb_columns_combined_layout(grid, header_idx)
            return (read(grid, aos_col), read(grid, consular_col))
        finally:
            wb.close()

    # ──────────────────────────────────────────────
    # Public API
    # ──────────────────────────────────────────────

    def get_eb_splits(self) -> list[EBSplit]:
        """National EB counts by fiscal year, split into AOS and consular."""
        if self._cache is None:
            self._cache = {}
            for fy, path in sorted(self._discover_files().items()):
                aos, consular = self._extract(path)
                self._cache[fy] = EBSplit(fy, aos, consular, path.name)
        return [self._cache[fy] for fy in sorted(self._cache)]

    def get_eb_splits_for_country(self, country: str) -> list[EBSplit]:
        """EB counts by fiscal year for one country of birth (e.g. 'India')."""
        results = []
        for fy, path in sorted(self._discover_files().items()):
            aos, consular = self._extract(path, country=country)
            results.append(EBSplit(fy, aos, consular, path.name))
        return results

    def get_coverage(self) -> list[int]:
        """Fiscal years for which a tables8-11 workbook is present."""
        return sorted(self._discover_files())
