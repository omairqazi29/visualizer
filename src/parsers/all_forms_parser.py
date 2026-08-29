"""Parser for the USCIS service-wide quarterly "All Forms" performance report.

File: data/USCIS_I485/quarterly_all_forms_fy####_q#_v#.xlsx

One row per USCIS form, grouped by category (Family Based, Employment Based,
Humanitarian, ...), with receipts, approvals, denials, completions, pending,
and median processing time for the quarter.

Why it is useful here: `i485_parser.py` covers I-485 flow and
`i140_radp_parser.py` covers I-140 detail, but neither gives the surrounding
context: the total agency workload, and the pending queue for the *other* forms
on the EB path (I-129, I-526, I-765, I-131). Processing time per form comes
straight from this report rather than being inferred.

Layout: a title block, then a header row naming the metric columns, then form
rows interleaved with category header rows (a category row has a label but no
form number and no counts).
"""

from pathlib import Path
from typing import Optional
import re

import openpyxl

from ..data_discovery import find_latest


__all__ = ["AllFormsParser"]


# Header captions mapped to the canonical field names this parser emits.
_COLUMN_ALIASES = {
    "forms received": "received",
    "received": "received",
    "approved": "approved",
    "denied": "denied",
    "total completions": "completions",
    "completions": "completions",
    "pending": "pending",
    "processing time": "processing_time_months",
}

_FORM_LABEL = "category and form number"
_FORM_TITLE = "form title"

_TOTAL_LABELS = {"total", "grand total"}
_SUPPRESSED = {"-", "D", "N/A", "NA", "X", ""}

# A form number looks like I-140, N-400, I-601A, I-129F.
_FORM_NUMBER = re.compile(r"^[A-Z]-?\d+[A-Z]?$", re.IGNORECASE)


def _clean(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).casefold()


def _strip_footnote(caption: str) -> str:
    """Drop the trailing footnote digits USCIS appends to header captions.

    e.g. 'Forms Received1' -> 'forms received', 'Processing Time6' -> 'processing time'.
    """
    return re.sub(r"\d+$", "", _clean(caption)).strip()


def _to_number(value, allow_float: bool = False):
    """Convert a count/time cell, or None when suppressed or unparseable."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if allow_float else int(round(float(value)))
    text = str(value).strip().replace(",", "")
    if text.upper() in _SUPPRESSED:
        return None
    try:
        return float(text) if allow_float else int(round(float(text)))
    except ValueError:
        return None


def get_latest_all_forms_path(data_dir: str = "data/USCIS_I485") -> Optional[str]:
    """Newest quarterly all-forms workbook, or None when none is present."""
    p = find_latest("quarterly_all_forms*.xlsx", data_dir)
    return str(p) if p is not None else None


class AllFormsParser:
    """Reads the USCIS service-wide quarterly all-forms report."""

    def __init__(self, path: str):
        self.path = Path(path)
        self._rows: Optional[list[dict]] = None
        self._period: Optional[str] = None

    @classmethod
    def latest(cls, data_dir: str = "data/USCIS_I485") -> Optional["AllFormsParser"]:
        """Parser for the newest all-forms file under data_dir, or None if absent."""
        path = get_latest_all_forms_path(data_dir)
        return cls(path) if path else None

    def _load(self) -> list[dict]:
        if self._rows is not None:
            return self._rows

        # An unreadable or truncated workbook degrades to "no data" rather than
        # raising: these arrive via the automated scanner, and a partial download
        # should not fail the request.
        try:
            wb = openpyxl.load_workbook(self.path, read_only=True, data_only=True)
        except Exception:
            self._rows = []
            return self._rows
        try:
            # The single data sheet is named per quarter, e.g. "FY26Q1_All_Forms"
            sheet = next((s for s in wb.sheetnames if "form" in _clean(s)), wb.sheetnames[0])
            grid = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
        except Exception:
            grid = []
        finally:
            wb.close()

        if not grid:
            self._rows = []
            return self._rows

        # Title block: the reporting period is the row naming a date range
        for row in grid[:6]:
            if row and row[0] and re.search(r"\d{4}", str(row[0])) and "-" in str(row[0]):
                self._period = " ".join(str(row[0]).split())
                break

        header_idx = next(
            (i for i, r in enumerate(grid) if r and _FORM_LABEL in _clean(r[0])),
            None,
        )
        if header_idx is None:
            self._rows = []
            return self._rows

        header = grid[header_idx]
        col_map: dict[int, str] = {}
        for col, cell in enumerate(header):
            field = _COLUMN_ALIASES.get(_strip_footnote(cell))
            if field and field not in col_map.values():
                col_map[col] = field

        title_col = next(
            (c for c, cell in enumerate(header) if _clean(cell) == _FORM_TITLE),
            1,
        )

        rows: list[dict] = []
        category = None
        for raw in grid[header_idx + 1:]:
            if not raw or not raw[0]:
                continue
            label = " ".join(str(raw[0]).split())
            key = _clean(label)

            values = {
                field: _to_number(raw[col], allow_float=(field == "processing_time_months"))
                for col, field in col_map.items()
                if col < len(raw)
            }
            has_counts = any(v is not None for f, v in values.items() if f != "processing_time_months")

            # A category header carries a label but no counts and no form number
            if not has_counts and not _FORM_NUMBER.match(label):
                if key not in _TOTAL_LABELS:
                    category = label
                continue

            record = {
                "form": label,
                "form_title": (
                    " ".join(str(raw[title_col]).split())
                    if title_col < len(raw) and raw[title_col]
                    else None
                ),
                "category": "TOTAL" if key in _TOTAL_LABELS else category,
                **values,
            }
            rows.append(record)

        self._rows = rows
        return self._rows

    # ──────────────────────────────────────────────
    # Public API
    # ──────────────────────────────────────────────

    @property
    def period(self) -> Optional[str]:
        """Reporting period, e.g. 'October 1, 2025 - December 31, 2025'."""
        self._load()
        return self._period

    def get_all_forms(self) -> list[dict]:
        """Every form row in the report."""
        return list(self._load())

    def get_by_category(self, category: str) -> list[dict]:
        """Form rows within one category, e.g. 'Employment Based'."""
        target = _clean(category)
        return [r for r in self._load() if _clean(r.get("category") or "") == target]

    def get_form(self, form: str) -> Optional[dict]:
        """One form's row, e.g. 'I-485'. Returns the first match.

        Some forms (notably I-130) appear more than once because USCIS reports
        separate rows per filing channel; use get_forms() for all of them.
        """
        matches = self.get_forms(form)
        return matches[0] if matches else None

    def get_forms(self, form: str) -> list[dict]:
        """All rows for one form number."""
        target = _clean(form)
        return [r for r in self._load() if _clean(r["form"]) == target]

    def get_totals(self) -> Optional[dict]:
        """The service-wide TOTAL row."""
        for r in self._load():
            if r.get("category") == "TOTAL":
                return r
        return None

    def get_eb_path_forms(self) -> list[dict]:
        """The forms on the employment-based green card path.

        Covers the petition stage (I-140, I-129, I-526), the adjustment stage
        (I-485), and the ancillary benefits filed with it (I-765 work
        authorization, I-131 advance parole), wherever each appears in the report.
        """
        wanted = ["I-140", "I-129", "I-526", "I-485", "I-765", "I-131"]
        found: list[dict] = []
        for form in wanted:
            found.extend(self.get_forms(form))
        return found
