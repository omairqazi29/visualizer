"""Parser for USCIS Monthly I-485 Receipts and Approvals data.

Tracks the I-485 queue flow: how many new applications are filed (receipts)
vs. how many are processed (approvals + denials) each month or quarter.
This models whether the I-485 queue is growing or shrinking over time.

Data sources:
- Monthly CSV reports (Congressional mandate):
  Published monthly by USCIS as "Number of Service-Wide Forms by Month,
  Form Status, and Processing Time."
  Files: data/USCIS_I485/monthly_{month}_{year}.csv
  Coverage: Jul 2024 – Feb 2026

- Quarterly XLSX performance data:
  Published quarterly by USCIS as "I-485 Performance Data by State."
  Files: data/USCIS_I485/i485_performance*_fy{year}_q{quarter}.xlsx
  Coverage: FY2024 Q1 – FY2025 Q4

Key notes:
- Monthly CSVs break down I-485 into sub-categories: Employment, Family,
  Asylum, Cuban, Refugee, Other.
- Quarterly XLSX files aggregate by Family-based, Employment-based,
  Humanitarian-based, and Others.
- Net flow = receipts − approvals − denials. Positive = queue growing.
- Numbers in CSVs have commas (e.g. "42,780") that must be stripped.
"""

import csv
import glob
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl


__all__ = ["I485FlowParser"]

# Month name → number mapping
_MONTH_MAP = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}

# Category extraction from I-485 description parenthetical
_CATEGORY_RE = re.compile(
    r"Application to Register Permanent Residence or Adjust Status \((\w+)\)"
)


def _parse_int(value) -> int:
    """Parse a comma-formatted integer string (e.g. '42,780' → 42780).

    Returns 0 for empty/unparseable values.
    """
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    cleaned = str(value).strip().replace(",", "")
    if not cleaned:
        return 0
    try:
        return int(cleaned)
    except ValueError:
        return 0


class I485FlowParser:
    """Parser for USCIS I-485 receipts, approvals, and pending data.

    Reads monthly CSVs and quarterly XLSX files from data/USCIS_I485/
    to produce time series of I-485 queue flow metrics.
    """

    def __init__(self, data_dir: str = "data/USCIS_I485"):
        self.data_dir = Path(data_dir)
        self._monthly_cache: Optional[list[dict]] = None
        self._quarterly_cache: Optional[list[dict]] = None

    # ──────────────────────────────────────────────
    # Monthly CSV loading
    # ──────────────────────────────────────────────

    def _load_monthly_csvs(self) -> list[dict]:
        """Load and parse all monthly CSV files.

        Each CSV has 4 header rows followed by column names and data.
        Filters for I-485 rows and extracts sub-category breakdowns
        (Employment, Family, Asylum, Cuban, Refugee, Other).

        Returns sorted list of dicts with flow metrics per month.
        """
        if self._monthly_cache is not None:
            return self._monthly_cache

        results = []
        pattern = str(self.data_dir / "monthly_*.csv")
        files = sorted(glob.glob(pattern))

        if not files:
            self._monthly_cache = []
            return self._monthly_cache

        for filepath in files:
            try:
                record = self._parse_monthly_csv(filepath)
                if record:
                    results.append(record)
            except Exception:
                # Skip corrupt / unreadable files
                continue

        # Sort chronologically by (year, month)
        results.sort(key=lambda r: (r["year"], r["month"]))
        self._monthly_cache = results
        return self._monthly_cache

    def _parse_monthly_csv(self, filepath: str) -> Optional[dict]:
        """Parse a single monthly CSV file into a flow record."""
        path = Path(filepath)

        # Extract month/year from filename: monthly_{month}_{year}.csv
        match = re.match(r"monthly_(\w+)_(\d{4})\.csv", path.name)
        if not match:
            return None
        month_name, year_str = match.group(1), match.group(2)
        month_num = _MONTH_MAP.get(month_name.lower())
        if month_num is None:
            return None
        year = int(year_str)

        # Read CSV, skipping first 4 header rows (use utf-8-sig to handle BOM)
        with open(filepath, "r", encoding="utf-8-sig") as f:
            for _ in range(4):
                next(f, None)
            reader = csv.DictReader(f)

            categories = {}
            for row in reader:
                form = (row.get("Form Number") or "").strip()
                if form != "I-485":
                    continue
                desc = row.get("Description") or ""
                cat_match = _CATEGORY_RE.search(desc)
                if not cat_match:
                    continue
                cat = cat_match.group(1)
                categories[cat] = {
                    "receipts": _parse_int(row.get("Forms Received", "0")),
                    "approvals": _parse_int(row.get("Approvals", "0")),
                    "denials": _parse_int(row.get("Denials", "0")),
                    "pending": _parse_int(row.get("Pending", "0")),
                    "pending_over_6mo": _parse_int(
                        row.get("Pending Over 6 Months", "0")
                    ),
                }

        if not categories:
            return None

        # Aggregate by EB / FB / Total
        eb = categories.get("Employment", {})
        fb = categories.get("Family", {})

        eb_receipts = eb.get("receipts", 0)
        eb_approvals = eb.get("approvals", 0)
        eb_denials = eb.get("denials", 0)
        eb_pending = eb.get("pending", 0)

        fb_receipts = fb.get("receipts", 0)
        fb_approvals = fb.get("approvals", 0)

        total_receipts = sum(c["receipts"] for c in categories.values())
        total_approvals = sum(c["approvals"] for c in categories.values())
        total_denials = sum(c["denials"] for c in categories.values())
        total_pending = sum(c["pending"] for c in categories.values())

        return {
            "period": f"{year}-{month_num:02d}",
            "year": year,
            "month": month_num,
            "source": "monthly",
            "months_covered": 1,
            "eb_receipts": eb_receipts,
            "eb_approvals": eb_approvals,
            "eb_denials": eb_denials,
            "eb_pending": eb_pending,
            "fb_receipts": fb_receipts,
            "fb_approvals": fb_approvals,
            "total_receipts": total_receipts,
            "total_approvals": total_approvals,
            "total_denials": total_denials,
            "total_pending": total_pending,
            "eb_net_flow": eb_receipts - eb_approvals - eb_denials,
            "total_net_flow": total_receipts - total_approvals - total_denials,
            "categories": categories,
        }

    # ──────────────────────────────────────────────
    # Quarterly XLSX loading
    # ──────────────────────────────────────────────

    def _load_quarterly_xlsx(self) -> list[dict]:
        """Load and parse all quarterly XLSX performance data files.

        Each XLSX has an I485_by_State sheet with:
        - Row 3: period string (e.g. "October 1, 2024 - December 31, 2024")
        - Row 7: Total aggregate row with category breakdowns
          Cols 3-6: FB | 7-10: EB | 11-14: Humanitarian | 15-18: Others | 19-22: Total

        Returns sorted list of dicts with flow metrics per quarter.
        """
        if self._quarterly_cache is not None:
            return self._quarterly_cache

        results = []
        pattern = str(self.data_dir / "i485_performance*.xlsx")
        files = sorted(glob.glob(pattern))

        if not files:
            self._quarterly_cache = []
            return self._quarterly_cache

        for filepath in files:
            try:
                record = self._parse_quarterly_xlsx(filepath)
                if record:
                    results.append(record)
            except Exception:
                # Skip corrupt / unreadable files
                continue

        results.sort(key=lambda r: (r["year"], r["month"]))
        self._quarterly_cache = results
        return self._quarterly_cache

    def _parse_quarterly_xlsx(self, filepath: str) -> Optional[dict]:
        """Parse a single quarterly XLSX file into a flow record."""
        wb = openpyxl.load_workbook(filepath)
        try:
            if "I485_by_State" not in wb.sheetnames:
                return None

            ws = wb["I485_by_State"]

            # Row 3: period string like "October 1, 2024 - December 31, 2024"
            period_str = ws.cell(row=3, column=1).value
            if not period_str:
                return None

            period_match = re.match(
                r"(\w+ \d+, \d{4})\s*-\s*(\w+ \d+, \d{4})", period_str
            )
            if not period_match:
                return None

            period_start = datetime.strptime(period_match.group(1), "%B %d, %Y")
            period_end = datetime.strptime(period_match.group(2), "%B %d, %Y")

            # Determine FY and quarter from period start month
            # FY starts Oct 1: Q1=Oct-Dec, Q2=Jan-Mar, Q3=Apr-Jun, Q4=Jul-Sep
            start_month = period_start.month
            if start_month >= 10:
                fy = period_start.year + 1
                quarter = 1
            elif start_month >= 7:
                fy = period_start.year
                quarter = 4
            elif start_month >= 4:
                fy = period_start.year
                quarter = 3
            else:
                fy = period_start.year
                quarter = 2

            # Row 7 total aggregate row (1-indexed columns)
            def cell_int(col: int) -> int:
                val = ws.cell(row=7, column=col).value
                if val is None:
                    return 0
                try:
                    return int(val)
                except (ValueError, TypeError):
                    return 0

            fb_receipts = cell_int(3)
            fb_approvals = cell_int(4)
            fb_denials = cell_int(5)
            fb_pending = cell_int(6)

            eb_receipts = cell_int(7)
            eb_approvals = cell_int(8)
            eb_denials = cell_int(9)
            eb_pending = cell_int(10)

            humanitarian_receipts = cell_int(11)
            humanitarian_approvals = cell_int(12)
            humanitarian_denials = cell_int(13)
            humanitarian_pending = cell_int(14)

            others_receipts = cell_int(15)
            others_approvals = cell_int(16)
            others_denials = cell_int(17)
            others_pending = cell_int(18)

            total_receipts = cell_int(19)
            total_approvals = cell_int(20)
            total_denials = cell_int(21)
            total_pending = cell_int(22)

            # Use end-of-quarter month as representative
            end_month = period_end.month
            end_year = period_end.year

            return {
                "period": f"FY{fy}-Q{quarter}",
                "year": end_year,
                "month": end_month,
                "source": "quarterly",
                "months_covered": 3,
                "period_start": period_start.strftime("%Y-%m-%d"),
                "period_end": period_end.strftime("%Y-%m-%d"),
                "fiscal_year": fy,
                "quarter": quarter,
                "eb_receipts": eb_receipts,
                "eb_approvals": eb_approvals,
                "eb_denials": eb_denials,
                "eb_pending": eb_pending,
                "fb_receipts": fb_receipts,
                "fb_approvals": fb_approvals,
                "total_receipts": total_receipts,
                "total_approvals": total_approvals,
                "total_denials": total_denials,
                "total_pending": total_pending,
                "eb_net_flow": eb_receipts - eb_approvals - eb_denials,
                "total_net_flow": total_receipts - total_approvals - total_denials,
                "categories": {
                    "Family": {
                        "receipts": fb_receipts,
                        "approvals": fb_approvals,
                        "denials": fb_denials,
                        "pending": fb_pending,
                    },
                    "Employment": {
                        "receipts": eb_receipts,
                        "approvals": eb_approvals,
                        "denials": eb_denials,
                        "pending": eb_pending,
                    },
                    "Humanitarian": {
                        "receipts": humanitarian_receipts,
                        "approvals": humanitarian_approvals,
                        "denials": humanitarian_denials,
                        "pending": humanitarian_pending,
                    },
                    "Others": {
                        "receipts": others_receipts,
                        "approvals": others_approvals,
                        "denials": others_denials,
                        "pending": others_pending,
                    },
                },
            }
        finally:
            wb.close()

    # ──────────────────────────────────────────────
    # Public API
    # ──────────────────────────────────────────────

    def get_monthly_series(self) -> list[dict]:
        """Return the monthly I-485 flow time series, sorted chronologically."""
        return self._load_monthly_csvs()

    def get_quarterly_series(self) -> list[dict]:
        """Return the quarterly I-485 flow time series, sorted chronologically."""
        return self._load_quarterly_xlsx()

    def get_eb_summary(self) -> dict:
        """Return summary statistics for EB I-485 flow.

        Computes averages, trends, and queue direction from monthly data.
        Falls back to quarterly data if no monthly data is available.

        Returns dict with:
        - latest_period: Most recent data period
        - latest_eb_pending: Current EB pending count
        - latest_total_pending: Current total pending count
        - avg_monthly_eb_receipts: Average monthly EB receipts
        - avg_monthly_eb_approvals: Average monthly EB approvals
        - avg_monthly_eb_net_flow: Average monthly EB net flow
        - queue_trend: "growing" or "shrinking"
        - pending_trend_pct: % change in recent vs older pending
        - data_points: Number of data points used
        - coverage: Date range string
        - source: "monthly" or "quarterly"
        """
        monthly = self._load_monthly_csvs()

        if not monthly:
            quarterly = self._load_quarterly_xlsx()
            if not quarterly:
                return {
                    "error": "No I-485 data available",
                    "data_points": 0,
                    "source": "none",
                }
            return self._summarize_quarterly(quarterly)

        return self._summarize_monthly(monthly)

    def _summarize_monthly(self, data: list[dict]) -> dict:
        """Compute summary statistics from monthly data points."""
        n = len(data)
        latest = data[-1]

        avg_eb_receipts = sum(d["eb_receipts"] for d in data) / n
        avg_eb_approvals = sum(d["eb_approvals"] for d in data) / n
        avg_eb_net = sum(d["eb_net_flow"] for d in data) / n

        # Pending trend: compare recent half vs older half
        mid = n // 2
        if mid > 0:
            older_pending = sum(d["eb_pending"] for d in data[:mid]) / mid
            recent_pending = sum(d["eb_pending"] for d in data[mid:]) / (n - mid)
            if older_pending > 0:
                pending_trend_pct = round(
                    (recent_pending - older_pending) / older_pending * 100, 1
                )
            else:
                pending_trend_pct = 0.0
        else:
            pending_trend_pct = 0.0

        queue_trend = "growing" if avg_eb_net > 0 else "shrinking"

        return {
            "latest_period": latest["period"],
            "latest_eb_pending": latest["eb_pending"],
            "latest_total_pending": latest["total_pending"],
            "avg_monthly_eb_receipts": round(avg_eb_receipts),
            "avg_monthly_eb_approvals": round(avg_eb_approvals),
            "avg_monthly_eb_net_flow": round(avg_eb_net),
            "queue_trend": queue_trend,
            "pending_trend_pct": pending_trend_pct,
            "data_points": n,
            "coverage": f"{data[0]['period']} to {latest['period']}",
            "source": "monthly",
        }

    def _summarize_quarterly(self, data: list[dict]) -> dict:
        """Compute summary statistics from quarterly data points.

        Converts quarterly totals to monthly averages (÷3) for comparability.
        """
        n = len(data)
        latest = data[-1]

        # Convert quarterly totals to monthly averages
        avg_eb_receipts = sum(d["eb_receipts"] for d in data) / (n * 3)
        avg_eb_approvals = sum(d["eb_approvals"] for d in data) / (n * 3)
        avg_eb_net = sum(d["eb_net_flow"] for d in data) / (n * 3)

        mid = n // 2
        if mid > 0:
            older_pending = sum(d["eb_pending"] for d in data[:mid]) / mid
            recent_pending = sum(d["eb_pending"] for d in data[mid:]) / (n - mid)
            if older_pending > 0:
                pending_trend_pct = round(
                    (recent_pending - older_pending) / older_pending * 100, 1
                )
            else:
                pending_trend_pct = 0.0
        else:
            pending_trend_pct = 0.0

        queue_trend = "growing" if avg_eb_net > 0 else "shrinking"

        return {
            "latest_period": latest["period"],
            "latest_eb_pending": latest["eb_pending"],
            "latest_total_pending": latest["total_pending"],
            "avg_monthly_eb_receipts": round(avg_eb_receipts),
            "avg_monthly_eb_approvals": round(avg_eb_approvals),
            "avg_monthly_eb_net_flow": round(avg_eb_net),
            "queue_trend": queue_trend,
            "pending_trend_pct": pending_trend_pct,
            "data_points": n,
            "coverage": f"{data[0]['period']} to {latest['period']}",
            "source": "quarterly",
        }
