"""Parser for the USCIS I-140 RADP report (Receipts, Approvals, Denials, Pending).

File: data/i140_fy####_q#_v#.xlsx

This is a different report from the two I-140 files already parsed here:

- `pipeline_parser.py` reads the I-140 *performance* file (backlog by country).
- `i140_receipts_parser.py` reads i140_rec_by_class_country (receipts by class).
- This one adds the full RADP flow: how many petitions were received, approved,
  denied, and are still **pending**, broken out by EB subcategory and by
  beneficiary country of birth, per fiscal quarter.

The pending counts are the notable addition: they are the live I-140 queue that
feeds future I-485 demand, at a granularity (EB1 vs EB2-NIW vs EB3) that the
aggregate performance file does not expose.

Sheets:
  RADP Summary  category totals per quarter (received/approved/denied/pending)
  Rec-COB       receipts by beneficiary country of birth
  App-COB       approvals by beneficiary country of birth
  Rec-State     receipts by beneficiary state
  App-State     approvals by beneficiary state

Layout notes:
- The country/state sheets use a two-row header: a group row spanning the EB
  preference ("First Preference (EB1)", ...) and a row naming each subcategory
  column. Which of those two rows carries the row-label caption varies between
  the Rec and App sheets, so both are located by content rather than position.
- Subcategory captions differ between the Rec and App sheets for the same
  column (e.g. "Professionals with Advanced Degrees (E21)" vs "Mem of
  Profession w/Adv Deg"). Preference-level totals are therefore derived from
  the group spans, not from caption matching.
"""

from pathlib import Path
from typing import Optional
import re

import openpyxl

from ..data_discovery import find_latest


__all__ = ["I140RADPParser"]


# Group headers that delimit the preference blocks on the country/state sheets
_PREFERENCE_GROUPS = {
    "first preference": "EB1",
    "second preference": "EB2",
    "third preference": "EB3",
}

# Metric columns on the RADP Summary sheet, in their published order
_SUMMARY_METRICS = ["received", "approved", "denied", "pending"]

# Summary row captions that are preference totals rather than subcategories
_SUMMARY_PREFERENCE_ROWS = {
    "first preference (eb1)": "EB1",
    "second preference (eb2)": "EB2",
    "third preference (eb3)": "EB3",
}

_TOTAL_LABELS = {"total", "grand total"}

# Matches a quarter block header ("1st Quarter", "2nd Quarter", ...). Anchored
# so the sheet subtitle "By Fiscal Year, Quarter, and Case Status" cannot be
# mistaken for a block header.
_QUARTER_HEADER = re.compile(r"^\d(?:st|nd|rd|th)\s+quarter$")

# USCIS suppression / placeholder markers
_SUPPRESSED = {"-", "D", "N/A", "NA", "X", ""}

# Everything below this caption on the summary sheet is footnotes, not data
_FOOTNOTE_MARKERS = ("table key", "references", "notes", "source")


def _clean(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).casefold()


def _to_int(value) -> Optional[int]:
    """Convert a count cell to int, or None when suppressed/unparseable."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(round(float(value)))
    text = str(value).strip().replace(",", "")
    if text.upper() in _SUPPRESSED:
        return None
    try:
        return int(round(float(text)))
    except ValueError:
        return None


def _visa_code(caption: str) -> Optional[str]:
    """Pull the visa code out of a subcategory caption, e.g. '... (E11)' -> 'E11'."""
    m = re.search(r"\(([EW]\w\d)\)", caption or "", re.IGNORECASE)
    if m:
        return m.group(1).upper()
    if "national interest waiver" in _clean(caption) or "niw" in _clean(caption):
        return "NIW"
    return None


def get_latest_radp_path(data_dir: str = "data") -> Optional[str]:
    """Newest RADP workbook, or None when none is present.

    Matched on 'i140_fy*' so this never collides with the differently-shaped
    i140_rec_by_class_country_* or eb_i140_* files in the same directory.
    """
    p = find_latest("i140_fy*.xlsx", data_dir)
    return str(p) if p is not None else None


class I140RADPParser:
    """Reads the USCIS I-140 RADP workbook."""

    def __init__(self, path: str):
        self.path = Path(path)
        self._grids: dict[str, list[list]] = {}

    @classmethod
    def latest(cls, data_dir: str = "data") -> Optional["I140RADPParser"]:
        """Parser for the newest RADP file under data_dir, or None if absent."""
        path = get_latest_radp_path(data_dir)
        return cls(path) if path else None

    # ──────────────────────────────────────────────
    # Sheet access
    # ──────────────────────────────────────────────

    def _grid(self, sheet: str) -> list[list]:
        """Materialize a sheet once (read-only sheets are forward-only streams).

        An unreadable or truncated workbook yields an empty grid rather than
        raising: these files arrive via the automated scanner, and a partial
        download should degrade to "no data" the way the other parsers here do,
        not fail the request.
        """
        if sheet not in self._grids:
            try:
                wb = openpyxl.load_workbook(self.path, read_only=True, data_only=True)
            except Exception:
                self._grids[sheet] = []
                return self._grids[sheet]
            try:
                match = next((s for s in wb.sheetnames if _clean(s) == _clean(sheet)), None)
                if match is None:
                    self._grids[sheet] = []
                else:
                    self._grids[sheet] = [list(r) for r in wb[match].iter_rows(values_only=True)]
            except Exception:
                self._grids[sheet] = []
            finally:
                wb.close()
        return self._grids[sheet]

    @property
    def period(self) -> Optional[str]:
        """The fiscal period the workbook covers, e.g. 'Fiscal Year 2026 (Q1-Q4)'.

        Requires a four-digit year in the caption so the sheet subtitle
        ("By Fiscal Year, Quarter, and Case Status") is not mistaken for it.
        """
        for row in self._grid("RADP Summary")[:6]:
            if not row or not row[0]:
                continue
            text = " ".join(str(row[0]).split())
            if "fiscal year" in _clean(text) and re.search(r"\d{4}", text):
                return text
        return None

    # ──────────────────────────────────────────────
    # Summary sheet
    # ──────────────────────────────────────────────

    def get_summary(self) -> dict:
        """Per-quarter category flows from the RADP Summary sheet.

        Returns {quarter_label: {category_key: {received, approved, denied, pending}}}
        where category_key is TOTAL, a preference (EB1/EB2/EB3), or a visa code
        (E11, E12, E13, E21, NIW, E31, E32, EW3).

        Quarters that have not been reported yet publish as all-zero blocks and
        are omitted.
        """
        grid = self._grid("RADP Summary")
        if not grid:
            return {}

        # Row 4 (index 3) labels each quarter block; each block is 4 columns
        quarter_row = next(
            (r for r in grid[:8] if r and any(_QUARTER_HEADER.match(_clean(c)) for c in r)),
            None,
        )
        if quarter_row is None:
            return {}
        blocks: list[tuple[str, int]] = [
            (" ".join(str(cell).split()), col)
            for col, cell in enumerate(quarter_row)
            if _QUARTER_HEADER.match(_clean(cell))
        ]
        if not blocks:
            return {}

        result: dict[str, dict] = {label: {} for label, _ in blocks}
        for row in grid:
            if not row or not row[0]:
                continue
            caption = _clean(row[0])
            if any(caption.startswith(m) for m in _FOOTNOTE_MARKERS):
                break
            if caption in _TOTAL_LABELS:
                key = "TOTAL"
            elif caption in _SUMMARY_PREFERENCE_ROWS:
                key = _SUMMARY_PREFERENCE_ROWS[caption]
            else:
                key = _visa_code(str(row[0]))
                if key is None:
                    continue

            for label, start in blocks:
                values = {}
                for offset, metric in enumerate(_SUMMARY_METRICS):
                    col = start + offset
                    values[metric] = _to_int(row[col]) if col < len(row) else None
                # Unreported quarters publish as all zeros
                if any(v for v in values.values()):
                    result[label][key] = values

        return {label: data for label, data in result.items() if data}

    # ──────────────────────────────────────────────
    # Country / state sheets
    # ──────────────────────────────────────────────

    @staticmethod
    def _locate_headers(grid: list[list]) -> tuple[Optional[int], Optional[int]]:
        """Return (group_row_idx, caption_row_idx) for a country/state sheet."""
        for i, row in enumerate(grid):
            if row and any("first preference" in _clean(c) for c in row):
                return i, i + 1
        return None, None

    @classmethod
    def _group_spans(cls, group_row: list, caption_row: list) -> dict[str, tuple[int, int]]:
        """Map preference -> (start_col, end_col_exclusive) from the group header.

        Merged header cells read as None, so each group runs until the next
        group label starts. The final group is additionally bounded by the
        captions: these sheets end with a published row-total column, which
        would otherwise be summed into the last preference and roughly double it.
        """
        starts: list[tuple[int, str]] = []
        for col, cell in enumerate(group_row):
            text = _clean(cell)
            for marker, pref in _PREFERENCE_GROUPS.items():
                if text.startswith(marker):
                    starts.append((col, pref))
                    break
        starts.sort()

        # Last data column = last captioned column that is not the row total
        last_data_col = -1
        for col, cell in enumerate(caption_row):
            caption = _clean(cell)
            if caption and caption not in _TOTAL_LABELS:
                last_data_col = col

        spans: dict[str, tuple[int, int]] = {}
        for i, (start, pref) in enumerate(starts):
            if i + 1 < len(starts):
                end = starts[i + 1][0]
            else:
                end = last_data_col + 1 if last_data_col >= start else len(group_row)
            spans[pref] = (start, end)
        return spans

    def _by_row_label(self, sheet: str) -> dict[str, dict]:
        """Parse a country/state sheet into {label: {EB1:.., EB2:.., EB3:.., detail:{}}}."""
        grid = self._grid(sheet)
        if not grid:
            return {}

        group_idx, caption_idx = self._locate_headers(grid)
        if group_idx is None:
            return {}

        group_row = grid[group_idx]
        caption_row = grid[caption_idx] if caption_idx < len(grid) else []
        spans = self._group_spans(group_row, caption_row)
        if not spans:
            return {}

        results: dict[str, dict] = {}
        # Data begins at the Grand Total row and runs to the end of the block
        started = False
        for row in grid[caption_idx + 1:]:
            if not row or not row[0]:
                if started:
                    break
                continue
            label = " ".join(str(row[0]).split())
            key = _clean(label)
            if not started:
                if key not in _TOTAL_LABELS:
                    continue
                started = True
            if any(key.startswith(m) for m in _FOOTNOTE_MARKERS):
                break

            entry: dict = {"detail": {}}
            for pref, (start, end) in spans.items():
                total = 0
                seen = False
                for col in range(start, min(end, len(row))):
                    value = _to_int(row[col])
                    if value is None:
                        continue
                    seen = True
                    total += value
                    caption = caption_row[col] if col < len(caption_row) else None
                    code = _visa_code(str(caption)) or f"{pref}_col{col - start}"
                    entry["detail"][code] = value
                entry[pref] = total if seen else None
            results[label] = entry

        return results

    def get_receipts_by_country(self) -> dict[str, dict]:
        """I-140 receipts by beneficiary country of birth."""
        return self._by_row_label("Rec-COB")

    def get_approvals_by_country(self) -> dict[str, dict]:
        """I-140 approvals by beneficiary country of birth."""
        return self._by_row_label("App-COB")

    def get_receipts_by_state(self) -> dict[str, dict]:
        """I-140 receipts by beneficiary state."""
        return self._by_row_label("Rec-State")

    def get_approvals_by_state(self) -> dict[str, dict]:
        """I-140 approvals by beneficiary state."""
        return self._by_row_label("App-State")

    def get_country_flow(self, country: str = "India") -> dict:
        """Receipts and approvals for one country of birth, by preference.

        Country labels in this report are upper-case (e.g. 'INDIA'); matching is
        case-insensitive.
        """
        target = _clean(country)

        def pick(table: dict[str, dict]) -> dict:
            for label, entry in table.items():
                if _clean(label) == target:
                    return entry
            return {}

        return {
            "country": country,
            "period": self.period,
            "receipts": pick(self.get_receipts_by_country()),
            "approvals": pick(self.get_approvals_by_country()),
        }
